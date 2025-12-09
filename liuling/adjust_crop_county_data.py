import logging
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook

from config import Configuration

# 设置日志
logging.basicConfig(
    filename="./adjust_crop_county_data.log",
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(name)s -   %(message)s",
    datefmt="%Y/%m/%d %H:%M:%S",
)
logger = logging.getLogger(__name__)

# --- 核心列名常量 ---
COL_NAME_CITY = 'City'  # 2022/2017数据中的城市列名
COL_NAME_CITY_CN = 'City_CN'
COL_NAME_COUNTY = 'County'  # 区县列名
COL_NAME_PROVINCE = 'Province'

# 定义需要处理"省统计年鉴"表的列名
COLS_CITY = [
    # 'Province',
    # 'City'
    'nonbeans_sown_area(1000 hectares)',
    'millet_sown_area(1000 hectares)',
    'sorghum_sown_area(1000 hectares)',
    'othercereals_sown_area(1000 hectares)',
    'potato_sown_area(1000 hectares)',
    'peanut_sown_area(1000 hectares)',
    'rapeseed_sown_area(1000 hectares)',
    'cotton_sown_area(1000 hectares)',
    'flax_sown_area(1000 hectares)',
    'sugarcane_sown_area(1000 hectares)',
    'sugarbeet_sown_area(1000 hectares)',
    'tobacco_sown_area(1000 hectares)',
    'vegetable_sown_area(1000 hectares)',
    'fruittree_sown_area(1000 hectares)',
    'greenfodder_sown_area(1000 hectares)',
    'managedgrass_sown_area(1000 hectares)',
    'naturalgrass_sown_area(1000 hectares)',
]

# --- 辅助：中英文省份名称映射表 (保持不变) ---
PROVINCE_MAP = {
    'Beijing': '北京',
    'Tianjin': '天津',
    'Hebei': '河北',
    'Shanxi': '山西',
    'Inner Mongolia': '内蒙古',
    'Liaoning': '辽宁',
    'Jilin': '吉林',
    'Heilongjiang': '黑龙江',
    'Shanghai': '上海',
    'Jiangsu': '江苏',
    'Zhejiang': '浙江',
    'Anhui': '安徽',
    'Fujian': '福建',
    'Jiangxi': '江西',
    'Shandong': '山东',
    'Henan': '河南',
    'Hubei': '湖北',
    'Hunan': '湖南',
    'Guangdong': '广东',
    'Guangxi': '广西',
    'Hainan': '海南',
    'Chongqing': '重庆',
    'Sichuan': '四川',
    'Guizhou': '贵州',
    'Yunnan': '云南',
    'Tibet': '西藏',
    'Shaanxi': '陕西',
    'Gansu': '甘肃',
    'Qinghai': '青海',
    'Ningxia': '宁夏',
    'Xinjiang': '新疆',
    # 请在此处补充您数据中所有省份的映射
}

COL_CROP_NAME_MAP = {
    '花生': 'peanut_sown_area(1000 hectares)',
    '油菜籽': 'rapeseed_sown_area(1000 hectares)',
    '棉花': 'cotton_sown_area(1000 hectares)',
    '麻类': 'flax_sown_area(1000 hectares)',
    '甘蔗': 'sugarcane_sown_area(1000 hectares)',
    '甜菜': 'sugarbeet_sown_area(1000 hectares)',
    '烟叶': 'tobacco_sown_area(1000 hectares)',
    '蔬菜': 'vegetable_sown_area(1000 hectares)',
    '果树': 'fruittree_sown_area(1000 hectares)',
}

# --- 关键修改：作物列名映射表 ---
# 格式：{ '2022数据中的列名': ('2017数据中的列名', '统计年鉴中的列名') }
# 注意：2022数据中的列名可能包含单位 '(1000 hectares)'
CROP_COLUMN_MAP = {
    # # 'nonbeans_sown_area(1000 hectares)': ('nonbeans_sown_area', ''),
    # 'millet_sown_area(1000 hectares)': ('millet_sown_area', '小米'),
    # 'sorghum_sown_area(1000 hectares)': ('sorghum_sown_area', '高粱'),
    # 'othercereals_sown_area(1000 hectares)': ('othercereals_sown_area', '其他杂粮'),
    # 'potato_sown_area(1000 hectares)': ('potato_sown_area', '马铃薯'),
    # 'peanut_sown_area(1000 hectares)': ('peanut_sown_area', '花生'),
    # 'rapeseed_sown_area(1000 hectares)': ('rapeseed_sown_area', '油菜籽'),  # 假设要调整油菜籽
    # 'cotton_sown_area(1000 hectares)': ('cotton_sown_area', '棉花'),  # 假设要调整棉花
    # 'flax_sown_area(1000 hectares)': ('flax_sown_area', '麻类'),
    # 'sugarcane_sown_area(1000 hectares)': ('sugarcane_sown_area', '甘蔗'),
    # 'sugarbeet_sown_area(1000 hectares)': ('sugarbeet_sown_area', '甜菜'),
    # 'tobacco_sown_area(1000 hectares)': ('tobacco_sown_area', '烟叶'),
    # 'vegetable_sown_area(1000 hectares)': ('vegetable_sown_area', '蔬菜'),
    # 'fruittree_sown_area(1000 hectares)': ('fruittree_sown_area', '果树'),
    # 'greenfodder_sown_area(1000 hectares)': ('greenfodder_sown_area', '青饲料'),
    # 'managedgrass_sown_area(1000 hectares)': ('managedgrass_sown_area', '管理草地'),
    # 'naturalgrass_sown_area(1000 hectares)': ('naturalgrass_sown_area', '自然草地'),
    # 'rice_sown_area(1000 hectares)': ('rice_sown_area', '水稻'),
    # 'wheat_sown_area(1000 hectares)': ('wheat_sown_area', '小麦'),
    # 'maize_sown_area(1000 hectares)': ('maize_sown_area', '玉米'),
    # 'beans_sown_area(1000 hectares)': ('beans_sown_area', '豆类'),

    'Agricultural Fertilizer Application (in pure nutrient equivalent)(10,000 tons)': (
        'Agricultural Fertilizer Application (in pure nutrient equivalent)(10,000 tons)', '化肥总量'),
    'Nitrogen Fertilizer Application(10,000 tons)': ('Nitrogen Fertilizer Application(10,000 tons)', '氮肥'),
    'Phosphate Fertilizer Application(10,000 tons)': ('Phosphate Fertilizer Application(10,000 tons)', '磷肥'),
    'Potassium Fertilizer Application(10,000 tons)': ('Potassium Fertilizer Application(10,000 tons)', '钾肥'),
    'compound Application(10,000 tons) ': ('compound Application(10,000 tons) ', '复合肥'),
    'rice_yield(10,000 tons)': ('rice_yield(10,000 tons)', '水稻'),
    'wheat_yield(10,000 tons)': ('wheat_yield(10,000 tons)', '小麦'),
    'maize_yield(10,000 tons)': ('maize_yield(10,000 tons)', '玉米'),
    'millet_yield(10,000 tons)': ('millet_yield(10,000 tons)', '谷子'),
    'sorghum_yield(10,000 tons)': ('sorghum_yield(10,000 tons)', '高粱'),
    'othercereals_yield(10,000 tons)': ('othercereals_yield(10,000 tons)', '其他谷物'),
    'beans_yield(10,000 tons)': ('beans_yield(10,000 tons)', '豆类'),
    'potato_yield(10,000 tons)': ('potato_yield(10,000 tons)', '马铃薯'),
    'peanut_yield(10,000 tons)': ('peanut_yield(10,000 tons)', '花生'),
    'rapeseed_yield(10,000 tons)': ('rapeseed_yield(10,000 tons)', '油菜籽'),
    'cotton_yield(10,000 tons)': ('cotton_yield(10,000 tons)', '棉花'),
    'flax_yield(10,000 tons)': ('flax_yield(10,000 tons)', '麻类'),
    'sugarcane_yield(10,000 tons)': ('sugarcane_yield(10,000 tons)', '甘蔗'),
    'sugarbeet_yield(10,000 tons)': ('sugarbeet_yield(10,000 tons)', '甜菜'),
    'tobacco_yield(10,000 tons)': ('tobacco_yield(10,000 tons)', '烟叶'),
    'vegetable_yield(10,000 tons)': ('vegetable_yield(10,000 tons)', '蔬菜'),
    'fruittree_yield(10,000 tons)': ('fruittree_yield(10,000 tons)', '果树'),

    # 请根据您的实际需求，补充需要进行调整的作物列
}


def clean_numeric_col(df, col):
    """将指定列转换为数值类型，并将 NaN 填充为 0。"""
    if col in df.columns:
        df[col] = df[col].astype(str).str.replace(',', '', regex=False).str.strip()
        df[col] = pd.to_numeric(df[col], errors='coerce')
        df[col] = df[col].fillna(0)
    return df


def load_data():
    """加载所有数据文件，并尝试加载辽宁省统计年鉴。"""
    try:
        # 1. 加载区县数据
        df_2022 = pd.read_excel(input_file, sheet_name="2022-crop-sown area")
        # 排除掉在指定列中全部为 NaN 的行
        # subset=target_columns: 指定只检查这些列
        # how='all': 表示只有当这些列中的值“全部”为 NaN 时才删除该行
        # df_2022 = df_2022.dropna(subset=COLS_CITY, how='all')
        df_2022 = df_2022.dropna(subset=['County', 'City', 'Province'], how='all')

        df_2017 = pd.read_excel(input_file, sheet_name="2017-crop-sown area")

        # 2. 加载国家统计年鉴数据
        df_national_ref = pd.read_excel(input_file, sheet_name="国家统计年鉴小作物种植面积", header=0)
        df_national_ref.rename(columns={'省份': COL_NAME_PROVINCE}, inplace=True)

        # rename_cols = COL_CROP_NAME_MAP.copy()
        # rename_cols.update({'省份': COL_NAME_PROVINCE})
        # df_national_ref.rename(columns=rename_cols, inplace=True)

        # 3. 尝试加载各省统计年鉴
        try:
            df_province_ref = pd.read_excel(input_file, sheet_name='各省统计年鉴', header=0)
            df_province_ref.rename(columns={'地区': COL_NAME_CITY_CN, '省': COL_NAME_PROVINCE}, inplace=True)

            # rename_cols = COL_CROP_NAME_MAP.copy()
            # rename_cols.update({'地区': COL_NAME_CITY_CN})
            # df_province_ref.rename(columns=rename_cols, inplace=True)
            print(f"注意: 工作表 '辽宁省统计年鉴' 已找到，将用于辽宁省的市级参考。")
        except ValueError as e:
            print(f"注意: 工作表 '各省统计年鉴' 未找到。所有省份的市级参考值将通过分解国家级数据得到。")
            raise e

        # 4. 清洗数值列
        # 对所有潜在需要调整的作物列进行清洗
        all_crop_cols = (set(CROP_COLUMN_MAP.keys())
                         | set(col[0] for col in CROP_COLUMN_MAP.values())
                         | set(col[1] for col in CROP_COLUMN_MAP.values())
                         )

        for col in all_crop_cols:
            df_2022 = clean_numeric_col(df_2022, col)
            df_2017 = clean_numeric_col(df_2017, col)
            df_national_ref = clean_numeric_col(df_national_ref, col)
            df_province_ref = clean_numeric_col(df_province_ref, col)

        # 5. 确保 City/Province 列为字符串
        for df in [df_2022, df_2017]:
            df[COL_NAME_CITY] = df[COL_NAME_CITY].astype(str)
            df[COL_NAME_PROVINCE] = df[COL_NAME_PROVINCE].astype(str)

        return df_2022, df_2017, df_national_ref, df_province_ref

    except FileNotFoundError as e:
        print(f"错误: 文件未找到，请检查路径是否正确: {e.filename}")
        raise e
    except Exception as e:
        print(f"加载数据时发生错误: {e}")
        raise e


def generate_province_statistical_table():
    pass


def process_data_for_all_crops_and_provinces(df_2022, df_2017, df_national_ref, df_province_ref):
    """循环遍历所有作物和所有省份，执行调整逻辑。"""

    # 最终调整后的数据副本 (在每个作物循环开始时创建，以便在 df_2022 上进行操作)
    df_adjusted_all = df_2022.copy()
    all_city_adjustment_logs = []
    all_province_adjustment_logs = []

    # 获取所有需要处理的省份 (英文名)
    provinces_to_process = df_2022[COL_NAME_PROVINCE].unique()

    print(f"--- 发现 {len(CROP_COLUMN_MAP)} 种作物，{len(provinces_to_process)} 个省份需要处理 ---")

    # --- 外层循环：遍历所有作物 ---
    for crop_2022_col, (crop_2017_col, crop_ref_col) in CROP_COLUMN_MAP.items():
        crop_2022 = crop_2022_col
        crop_2017 = crop_2017_col
        crop_REF = crop_ref_col

        print(f"\n\n#######################################################")
        print(f"### 开始调整作物：{crop_REF} ({crop_2022}) ###")
        print(f"#######################################################")

        # --- 内层循环：遍历所有省份 ---
        for province_en in provinces_to_process:
            province_cn = PROVINCE_MAP.get(province_en)
            if not province_cn:
                print(f"\n警告: 未找到省份 '{province_en}' 的中文名称映射，跳过此省份调整。")
                continue

            print(f"\n=======================================================")
            print(f"开始处理 {crop_REF} 在省份: {province_cn} ({province_en}) 的数据")
            print(f"=======================================================")

            log_province_entry = revise_by_province(crop_2017, crop_2022, crop_REF,
                                                    df_2017, df_adjusted_all, df_national_ref, df_province_ref,
                                                    province_cn, province_en,
                                                    all_city_adjustment_logs)
            if log_province_entry is not None:
                all_province_adjustment_logs.append(log_province_entry)

    # --- 5. 结果汇总与输出 ---
    print("\n\n#######################################################")
    print("### 最终处理结果汇总：市级调整日志 (所有作物/所有省份) ###")
    print("#######################################################")

    df_city_log = pd.DataFrame(all_city_adjustment_logs)
    print(df_city_log.to_markdown(index=False))

    # 定义另一个样式函数：为特定字符串设置绿色背景和白色字体
    def highlight_status(val):
        if val is None or val is np.nan:
            return ''

        if 'Error' in val:
            # 设置背景为红色
            return 'background-color: red;'
        elif 'But 2017 base is zero or missing' in val:
            # 设置背景为黄色
            return 'background-color: yellow'
        elif 'Adjusted. New Sum' in val:
            return 'background-color: green'
        elif '范围内，进行市级调整' in val:
            return 'background-color: green'
        return ''

    # 应用样式到 Styler 对象
    df_styled_city_log = df_city_log.style.map(
        highlight_status,  # 应用到整个 DataFrame，检查负数
        subset=['Status']  # 仅应用于Status列
    )
    write_df(df_styled_city_log, output_log_file, sheet_name="城市统计年鉴校正日志")

    df_province_log = pd.DataFrame(all_province_adjustment_logs)
    df_styled_province_log = df_province_log.style.map(
        highlight_status,  # 应用到整个 DataFrame，检查负数
        subset=['Status']  # 仅应用于Status列
    )
    write_df(df_styled_province_log, output_log_file, sheet_name="省统计年鉴校正日志")

    # 最终检查并输出省份总和
    print(f"\n\n#######################################################")
    print(f"### 3. 最终省份/作物总和检查 ###")
    print(f"#######################################################")
    final_check(df_adjusted_all, df_national_ref, provinces_to_process)

    # 保存最终结果到 Excel
    df_adjusted_all = df_adjusted_all.dropna(subset=['County', 'City', 'Province'], how='all')
    df_adjusted_result = mask_changed(df_2022, df_adjusted_all)

    write_df(df_adjusted_result, output_adjusted_file)
    print(f"\n数据处理完成。调整后的数据已保存到文件: {output_adjusted_file}")

    return output_adjusted_file


def revise_by_province(crop_2017, crop_2022, crop_ref,
                       df_2017, df_adjusted_all, df_national_ref, df_province_ref,
                       province_cn, province_en,
                       all_city_adjustment_logs):
    if crop_2022 not in df_adjusted_all.columns:
        return None

    # 过滤出 2022 年当前省份的所有区县数据
    # 注意：此处使用 df_adjusted_all，确保上一省份/作物调整的结果得以保留
    df_province_data = df_adjusted_all[df_adjusted_all[COL_NAME_PROVINCE] == province_en].copy()

    # --- 2. 省份层面检查 (AO20/C8) ---
    AO20 = df_province_data[crop_2022].sum()
    df_2017_province = df_2017[df_2017[COL_NAME_PROVINCE] == province_en]
    AO20_2017 = df_2017_province[crop_2017].sum() if crop_2017 in df_2017_province else None

    C8_series = df_national_ref[df_national_ref[COL_NAME_PROVINCE] == province_cn][crop_ref] \
        if crop_ref in df_national_ref.columns else None
    # 国家统计年鉴中如果没有 则从对应省统计年鉴中找对应数据
    if C8_series is None or C8_series.empty:
        _df_province = df_province_ref[df_province_ref[COL_NAME_PROVINCE] == province_cn]
        if not _df_province.empty:
            C8_series = _df_province[crop_ref] if crop_ref in df_province_ref.columns else None

    C8 = C8_series.iloc[0] if C8_series is not None and not C8_series.empty else np.nan
    logger.info(f"1. 2022 {province_cn} {crop_ref}({crop_2022}) 面积 (区县总和 AO20): {AO20:.4f} 千公顷")
    logger.info(f"   国家统计年鉴 {province_cn} {crop_ref} 数据 (C8): {C8:.4f} 千公顷")

    ratio_province = AO20 / C8 if AO20 > 0 and C8 > 0 else 0

    log_province_entry = {
        'Crop   ': crop_ref,
        'Crop_2022 Col_Name': crop_2022,
        'Crop_2017 Col_Name': crop_2017,
        COL_NAME_PROVINCE + "   ": province_en,
        'County Province Sum(AO20)': AO20,
        '2017 County Province Sum(AO20)': AO20_2017,
        'National Province Ref(C8)': C8,
        'AO20/C8 Ratio': ratio_province,
        'Status': '',
    }

    needs_city_adjustment = True
    if not np.isnan(C8) and C8 > 0 and AO20 > 0:
        print(f"   AO20/C8 比例: {ratio_province:.4f}")

        if RATIO_MIN_PROVINCE <= ratio_province <= RATIO_MAX_PROVINCE:
            # AO20/C8结果范围在0.95-1.05之间 不需要调整
            print(f"   比例在 [{RATIO_MIN_PROVINCE}, {RATIO_MAX_PROVINCE}] 范围内，无需市级调整。")
            needs_city_adjustment = False

            log_province_entry[
                "Status"] = f'Skipped 【Ratio AO20/C8 = {ratio_province} in ({RATIO_MIN_PROVINCE}-{RATIO_MAX_PROVINCE})】'
            return log_province_entry
        else:
            # AO20/C8结果范围在0.95-1.05之外, 进行调整
            print(f"   比例不在 [{RATIO_MIN_PROVINCE}, {RATIO_MAX_PROVINCE}] 范围内，进行市级调整。")
            log_province_entry[
                "Status"] = f'比例不在 [{RATIO_MIN_PROVINCE}, {RATIO_MAX_PROVINCE}] 范围内，进行市级调整。'
    # else:
    #     print("   C8 或 AO20 数据无效，跳过省份级检查，无法进行市级调整。")
    #     needs_city_adjustment = False
    #     log_province_entry["Status"] = f'C8 或 AO20 数据无效，跳过省份级检查，无法进行市级调整。'
    # continue

    # --- 2.3 确定市级参考数据 (B) ---
    _df_city_ref = pd.DataFrame

    # 优先使用提供的各省统计年鉴
    if not df_province_ref.empty:
        # if CROP_REF in df_liaoning_ref.columns:
        _df_city_ref = df_province_ref.copy()
        print(f"--- 使用提供的 '{province_cn}统计年鉴' {crop_ref} 数据作为市级参考值 (B)。 ---")

    # 其他情况（包括辽宁年鉴缺失时）进行推算
    # elif needs_city_adjustment:
    #     # 推算逻辑：按 2022 区县数据中的城市占比分解 C8
    #     print(f"--- 缺乏市级年鉴数据，根据2022区县数据比例分解 C8 ({province_cn}) ---")
    #
    #     df_city_sums = df_province_data.groupby(COL_NAME_CITY)[crop_2022].sum().reset_index()
    #     df_city_sums['Weight'] = df_city_sums[crop_2022] / AO20
    #     df_city_sums[crop_REF] = df_city_sums['Weight'] * C8
    #     _df_city_ref = df_city_sums.rename(columns={COL_NAME_CITY: 'City_CN'})[['City_CN', crop_REF]]

    df_city_ref = _df_city_ref[_df_city_ref[COL_NAME_PROVINCE] == province_cn]
    # --- 3. 市级调整逻辑 (统一使用 df_city_ref) ---
    if needs_city_adjustment and not df_city_ref.empty:
        cities_to_process = [c for c in df_city_ref['City_CN'].unique() if
                             c not in ['全省', '地区', 'nan', np.nan, province_cn]]

        # --- 内层循环：遍历某个省下面的所有市 ---
        for city_cn in cities_to_process:
            log_city_entry = revise_by_city(crop_2022, crop_2017, crop_ref,
                                            df_city_ref, df_province_data, df_adjusted_all,
                                            province_en, city_cn)
            all_city_adjustment_logs.append(log_city_entry)

    else:
        print(f"   {province_cn} {crop_ref} 市级参考数据缺失或不需要调整，跳过市级循环。")

    # 修正数据后校验各省不同作物数据总和
    check_province(C8, crop_2022, province_en, df_adjusted_all, log_province_entry)

    return log_province_entry


def revise_by_city(crop_2022, crop_2017, crop_REF,
                   df_city_ref, df_province_data, df_adjusted_all,
                   province_en, city_cn):
    # 3.1 查找 B
    B_series = df_city_ref[df_city_ref['City_CN'] == city_cn][crop_REF] \
        if crop_REF in df_city_ref.columns else None
    B = B_series.iloc[0] if B_series is not None else 0.0

    # 确定 City 匹配方式
    # city_match_name = city_cn
    # city_match_name = city_cn.replace('市', '').replace('地区', '')
    df_city_data = df_province_data[df_province_data[COL_NAME_CITY] == city_cn].copy()

    A = 0
    city_name_in_data = None
    if not df_city_data.empty:
        city_name_in_data = df_city_data[COL_NAME_CITY].iloc[0]
        A = df_city_data[crop_2022].sum()

    df_city_2017_base = df_2017[
        (df_2017[COL_NAME_CITY] == city_name_in_data) &
        (df_2017[COL_NAME_PROVINCE] == province_en)
        ].copy()
    sum_2017 = df_city_2017_base[crop_2017].sum() if crop_2017 in df_city_2017_base else 0.0

    AB_ratio = A / B if B != 0 else 0.0
    log_city_entry = {
        'Crop': crop_REF,
        'Crop_2022': crop_2022,
        'Crop_2017': crop_2017,
        'Province': province_en,
        'City': city_name_in_data,
        '2022 County_Sum (A)': A,
        'Ref_Data (B)': B,
        '2017 County_Sum': sum_2017,
        'A/B Ratio': AB_ratio,
        'Status': ''
    }

    if df_city_data.empty or (A <= 0 and B <= 0 and sum_2017 <= 0):
        log_city_entry['Status'] = 'Skipped 【A and B and sum_2017 is zero/missing】'
        return log_city_entry

    # if B <= 0:
    #     """
    #     查询2022年表中数值为0的区县, 并对应这些区县找到2017年表中不为0的数据 填充到2022年的表格中
    #     """
    #     df_city_2022_zero = df_city_data[df_city_data[crop_2022] == 0].copy()
    #     counties_to_adjust = df_city_2022_zero[COL_NAME_COUNTY].tolist()
    #
    #     df_city_2017_non_zero = df_city_2017_base[
    #         (df_city_2017_base[crop_2017] != 0) &
    #         (df_city_2017_base[COL_NAME_COUNTY].isin(counties_to_adjust))
    #         ]
    #     if not df_city_2022_zero.empty and not df_city_2017_non_zero.empty:
    #         df_city_2022_zero['Adjustment'] = df_city_2017_non_zero[crop_2017]
    #         for _, row in df_city_2022_zero.iterrows():
    #             county = row[COL_NAME_COUNTY]
    #             adjustment = row['Adjustment']
    #             condition = (df_adjusted_all[COL_NAME_CITY] == city_name_in_data) & \
    #                         (df_adjusted_all[COL_NAME_COUNTY] == county) & \
    #                         (df_adjusted_all[COL_NAME_PROVINCE] == province_en)
    #             df_adjusted_all.loc[condition, crop_2022] = adjustment
    #     log_city_entry['Status'] = 'Replace 2022 data with 2017 data'

    # 3.3 比例判断与调整
    # 比例范围在0.99-1.01之间, 无需调整
    if RATIO_MIN_1 <= AB_ratio <= RATIO_MAX_1:
        log_city_entry['Status'] = f'No Adjustment 【Ratio({AB_ratio}) within 0.99-1.01】'

    # 比例范围在<0.9 或 >1.1 范围内
    elif AB_ratio < RATIO_MIN_2 or AB_ratio > RATIO_MAX_2:
        if B > 1.0:
            revise_by_where(AB_ratio, A, B, crop_2022, crop_2017,
                            df_city_data, df_adjusted_all,
                            city_name_in_data, province_en,
                            log_city_entry)
        else:
            if B < 1.0 and -0.5 < (A - B) < 0.5:
                log_city_entry[
                    'Status'] = f'Adjustment needed 【B({B})<1.0 and -0.5 < A-B({A - B}) < 0.5】'
                # --- 4. 执行调整 ---
                revise_county(A, B, crop_2022, crop_2017,
                              df_city_data, df_adjusted_all,
                              city_name_in_data, province_en,
                              log_city_entry)
            else:
                revise_by_where(AB_ratio, A, B, crop_2022, crop_2017,
                                df_city_data, df_adjusted_all,
                                city_name_in_data, province_en,
                                log_city_entry)

    # 比例范围在0.9-0.99，1.01-1.1之内, 进行如下调整
    elif RATIO_MIN_2 <= AB_ratio < RATIO_MIN_1 or RATIO_MAX_1 < AB_ratio <= RATIO_MAX_2:
        log_city_entry[
            'Status'] = f'Adjustment needed 【Ratio({AB_ratio}) in ({RATIO_MIN_2}-{RATIO_MIN_1}) or ({RATIO_MAX_1}-{RATIO_MAX_2})】'
        # --- 4. 执行调整 ---
        revise_county(A, B, crop_2022, crop_2017,
                      df_city_data, df_adjusted_all,
                      city_name_in_data, province_en,
                      log_city_entry)

    return log_city_entry


def revise_by_where(AB_ratio, A, B, crop_2022, crop_2017,
                    df_city_data, df_adjusted_all,
                    city_name_in_data, province_en,
                    log_city_entry):
    # 如果 A/B < 0.9, 执行步骤三调整
    if AB_ratio < RATIO_MIN_2:
        # --- 4. 执行调整 ---
        log_city_entry[
            'Status'] = f'Adjustment needed 【B({B})>1.0 and Ratio({AB_ratio}) < {RATIO_MIN_2}】'
        revise_county(A, B, crop_2022, crop_2017,
                      df_city_data, df_adjusted_all,
                      city_name_in_data, province_en,
                      log_city_entry)

    # 如果 A/B > 1.1, 按条件调整
    elif AB_ratio > RATIO_MAX_2:
        # #  如果 A/B > 1.1 && < 1.5
        # # if RATIO_MAX_3 > AB_ratio > RATIO_MAX_2:
        # if AB_ratio > RATIO_MAX_2:
        #     # 当2022年的数据中都有值则调整, 否则标记
        #     df_city_2022_zero = df_city_data[df_city_data[crop_2022] == 0]
        #     if df_city_2022_zero.empty:
        #         log_city_entry[
        #             'Status'] = f'Adjustment needed 【B({B})>1.0 and {RATIO_MAX_3} > Ratio({AB_ratio}) > {RATIO_MAX_2}】'
        #         revise_county(A, B, crop_2022, crop_2017,
        #                       df_city_data, df_adjusted_all,
        #                       city_name_in_data, province_en,
        #                       log_city_entry)
        #     else:
        #         log_city_entry[
        #             'Status'] = f'Marked 【Ratio({AB_ratio}) < {RATIO_MAX_3} and > {RATIO_MAX_2}, But 2022 base has zero'
        # #  如果 A/B >= 1.5 报错
        # elif AB_ratio >= RATIO_MAX_3:
        #     log_city_entry['Status'] = f'Error |【Ratio({AB_ratio}) >= {RATIO_MAX_3}】'

        log_city_entry['Status'] = f'Adjustment needed 【Ratio({AB_ratio}) >= {RATIO_MAX_3}】'
        revise_county(A, B, crop_2022, crop_2017,
                      df_city_data, df_adjusted_all,
                      city_name_in_data, province_en,
                      log_city_entry)


def revise_county(A, B, crop_2022, crop_2017,
                  df_city_data, df_adjusted_all,
                  city_name_in_data, province_en,
                  log_city_entry):
    # B为省统计年鉴中各市的总和
    if B == 0:
        log_city_entry['Status'] = f'Skipped 【A={A}, B={B}】'
        return

    difference = A - B
    if difference > 0:
        """
        计算（A-B) > 0, 做如下调整：
        按照2022年已有数据比例整体下调到B
        """
        df_city_2022_non_zero = df_city_data[df_city_data[crop_2022] != 0]
        counties_to_adjust = df_city_2022_non_zero[COL_NAME_COUNTY].tolist()

        df_city_2022_base = df_2022[
            (df_2022[COL_NAME_CITY] == city_name_in_data) &
            (df_2022[COL_NAME_PROVINCE] == province_en) &
            (df_2022[COL_NAME_COUNTY].isin(counties_to_adjust))
            ].copy()

        if not df_city_2022_base.empty:
            df_city_2022_base['Adjustment_Ratio'] = round(1 - difference / A, ROUND_DECIMAL)
            df_city_2022_base['Adjustment'] = round(
                df_city_2022_base[crop_2022] * df_city_2022_base['Adjustment_Ratio'], ROUND_DECIMAL)

            # 4. 更新 df_adjusted_all 中的数据
            for _, row in df_city_2022_base.iterrows():
                update_crop_data(df_adjusted_all, df_city_2022_base, B, city_name_in_data, crop_2022, province_en)
        else:
            log_city_entry['Status'] += f' | Adjustment Failed - 2022 base city is zero or missing.'
    else:
        """
        计算（A-B) < 0, 对2022年鞍山市花生数据向上调整，例如：
        选择2022年鞍山市花生数据为0或空白的区县对应选择“2017-crop-sown area”表格中2017年相同区县（假设有5个区县a,b,c,d,e 数据分别为1，2，3，4，5）；
        将（A-B）*1/（1+2+3+4+5）填入2022年鞍山市花生数据中的a区县
        """
        df_city_2022_zero = df_city_data[df_city_data[crop_2022] == 0]
        counties_to_adjust = df_city_2022_zero[COL_NAME_COUNTY].tolist()

        # 2. 匹配 2017 年数据 (注意使用 CROP_2017 列名)
        df_city_2017_base = df_2017[
            (df_2017[COL_NAME_CITY] == city_name_in_data) &
            (df_2017[COL_NAME_PROVINCE] == province_en) &
            (df_2017[COL_NAME_COUNTY].isin(counties_to_adjust))
            ].copy()

        sum_2017 = df_city_2017_base[crop_2017].sum() if crop_2017 in df_city_2017_base else 0.0
        log_city_entry['2017 County_Sum'] += sum_2017

        if sum_2017 > 0 and not df_city_2017_base.empty:
            df_city_2017_base['Adjustment_Ratio'] = round(df_city_2017_base[crop_2017] / sum_2017, ROUND_DECIMAL)
            df_city_2017_base['Adjustment'] = round(-difference * df_city_2017_base['Adjustment_Ratio'], ROUND_DECIMAL)

            # df_city_2017_base['Adjustment_Ratio'] = round(df_city_2017_base[crop_2017] / sum_2017, ROUND_DECIMAL)
            # if sum_2017 - B > 0:
            #     df_city_2017_base['Adjustment'] = round(-difference * df_city_2017_base['Adjustment_Ratio'], ROUND_DECIMAL)
            # else:
            #     df_city_2017_base['Adjustment'] = round(difference * df_city_2017_base['Adjustment_Ratio'], ROUND_DECIMAL)

            # 4. 更新 df_adjusted_all 中的数据
            update_crop_data(df_adjusted_all, df_city_2017_base, B, city_name_in_data, crop_2022, province_en)
        else:
            """
            如果2022年鞍山市花生数据有区县不为0的 则对不为0的数据向上调整
            """
            df_city_2022_non_zero = df_city_data[df_city_data[crop_2022] != 0]
            if not df_city_2022_non_zero.empty:  # 2022年中有不为0的数据
                counties_to_adjust = df_city_2022_non_zero[COL_NAME_COUNTY].tolist()

                df_city_2022_base = df_2022[
                    (df_2022[COL_NAME_CITY] == city_name_in_data) &
                    (df_2022[COL_NAME_PROVINCE] == province_en) &
                    (df_2022[COL_NAME_COUNTY].isin(counties_to_adjust))
                    ].copy()

                if not df_city_2022_base.empty:
                    # df_city_2022_base['Adjustment_Ratio'] = round(1 + -difference / B, ROUND_DECIMAL)
                    # df_city_2022_base['Adjustment'] = round(df_city_2022_base[crop_2022] * df_city_2022_base['Adjustment_Ratio'], ROUND_DECIMAL)

                    df_city_2022_base['Adjustment_Ratio'] = round(df_city_2022_base[crop_2022] / A, ROUND_DECIMAL)
                    df_city_2022_base['Adjustment'] = round(
                        df_city_2022_base[crop_2022] + (-difference * df_city_2022_base['Adjustment_Ratio']),
                        ROUND_DECIMAL)

                    # 4. 更新 df_adjusted_all 中的数据
                    update_crop_data(df_adjusted_all, df_city_2022_base, B, city_name_in_data, crop_2022, province_en)

                    # log_entry['Status'] += f' | Adjustment Failed - 2017 base SUM_2017({sum_2017}) is zero or missing.'
            else:
                if sum_2017 == 0:
                    log_city_entry['Status'] += f"| But 2017 base is zero or missing."
                    return

    # 5. 检查修改后与修改前的结果一致性
    check_city(df_adjusted_all, B, city_name_in_data, crop_2022, log_city_entry, province_en)


def update_crop_data(df_adjusted_all, df_city_base, B, city_name_in_data, crop_2022, province_en):
    for _, row in df_city_base.iterrows():
        county = row[COL_NAME_COUNTY]
        adjustment = row['Adjustment']

        condition = (df_adjusted_all[COL_NAME_CITY] == city_name_in_data) & \
                    (df_adjusted_all[COL_NAME_COUNTY] == county) & \
                    (df_adjusted_all[COL_NAME_PROVINCE] == province_en)

        # 核心：动态更新 CROP_2022 所在的列
        df_adjusted_all.loc[condition, crop_2022] = adjustment


def check_province(C8: float, crop_2022: str, province_en: str, df_adjusted_all, log_province_entry):
    if '范围内，进行市级调整' in log_province_entry['Status']:
        new_AO20 = df_adjusted_all[(df_adjusted_all[COL_NAME_PROVINCE] == province_en)][crop_2022].sum()
        new_AO20 = round(new_AO20, ROUND_DECIMAL)

        new_ratio_province = new_AO20 / C8 if new_AO20 > 0 and C8 > 0 else 0
        log_province_entry['Status'] += f' | Adjusted. New Sum: {new_AO20}'
        log_province_entry['Check (New AO20/C8 Ratio)'] = new_ratio_province


def check_city(df_adjusted_all, B, city_name_in_data, crop_2022, log_entry, province_en):
    A_new = df_adjusted_all[
        (df_adjusted_all[COL_NAME_CITY] == city_name_in_data) &
        (df_adjusted_all[COL_NAME_PROVINCE] == province_en)
        ][crop_2022].sum()

    A_new = round(A_new, ROUND_DECIMAL)
    # B = round(B, ROUND_DECIMAL)
    new_AB_ratio = round(A_new / B, 4) if B != 0 else 0.0
    logger.info(
        f"{province_en} / {city_name_in_data} {crop_2022} | Check (New A/B) = ( {A_new} / {B} = {new_AB_ratio} )")

    log_entry['Check (New A/B)'] = f"{A_new} / {B} = {new_AB_ratio}"
    log_entry['Status'] += f" | Adjusted. New Sum: {A_new}"


def final_check(df_adjusted_all, df_national_ref, provinces_to_process):
    for crop_2022_col, (_, crop_ref_col) in CROP_COLUMN_MAP.items():
        crop_ref = crop_ref_col
        print(f"\n--- {crop_ref} 调整后总和 ---")
        for province_en in provinces_to_process:
            province_cn = PROVINCE_MAP.get(province_en)
            if not province_cn:
                continue

            df_adjusted_province = df_adjusted_all[df_adjusted_all[COL_NAME_PROVINCE] == province_en]
            AO20_new = df_adjusted_province[crop_2022_col].sum() \
                if crop_2022_col in df_adjusted_province.columns else 0.0

            df_national_province = df_national_ref[df_national_ref[COL_NAME_PROVINCE] == province_cn]
            C8_series = df_national_province[crop_ref] if crop_ref in df_national_province else pd.Series([])
            C8 = C8_series.iloc[0] if not C8_series.empty else np.nan

            if not np.isnan(C8) and C8 > 0:
                ratio_province_new = AO20_new / C8
                print(
                    f"{province_cn} 调整后总和: {AO20_new:.4f} 千公顷 | 调整后比例 (AO20/C8): {ratio_province_new:.4f}")
            else:
                print(f"{province_cn} 调整后总和: {AO20_new:.4f} 千公顷 | C8数据缺失。")


def mask_changed(df_2022, df_adjusted_all) -> pd.DataFrame:
    all_crop_cols = (set(CROP_COLUMN_MAP.keys())
                     | set(col[0] for col in CROP_COLUMN_MAP.values())
                     | set(col[1] for col in CROP_COLUMN_MAP.values())
                     )

    for col in all_crop_cols:
        df_adjusted_all = clean_numeric_col(df_adjusted_all, col)

    # 生成差异布尔掩码，并排除 NaN ---
    # 步骤 a: 找出所有不相等的值 (包括 NaN 引起的不相等)
    changed_mask = (df_adjusted_all != df_2022)
    # 步骤 b: 找出 df_new 中非 NaN 的值
    not_na_mask = df_adjusted_all.notna()
    # 步骤 c: 结合两个掩码：必须是不相等 (changed_mask) 且在新数据中不能是 NaN (not_na_mask)
    # 只有当一个值发生了变化 **并且** 这个新的值不是 NaN 时，才将它标记为 True
    final_mask = changed_mask & not_na_mask

    # 针对字符串/对象类型列进行更细致的 NaN 比较
    # 对于字符串，空字符串 '' 和 NaN 应该被视为不等，但 NaN 和 NaN 应该相等
    # for col in df_2022.select_dtypes(include=['object']).columns:
    #     mask_na_both = df_2022[col].isna() & df_adjusted_all[col].isna()
    #     mask_diff = (df_2022[col] != df_adjusted_all[col]) | (df_2022[col].isna() != df_adjusted_all[col].isna())
    #     final_mask[col] = mask_diff & ~mask_na_both

    # 应用样式
    # ⭐️ 关键：使用 style.apply(func, axis=None) 将整个 DataFrame 传递给函数
    df_adjusted_result = df_adjusted_all.style.apply(
        highlight_changed_cells,
        mask=final_mask,  # 将布尔掩码作为关键字参数传递
        axis=None,  # 确保函数作用于整个表格
        color='lightcoral'  # 使用更醒目的颜色进行标记
    )
    return df_adjusted_result


def highlight_changed_cells(df_data, mask, color='red'):
    """
    df_data: Styler.apply(axis=None) 传递的整个 df_new 数据。
    mask: 预先计算好的布尔掩码 (changed_mask)。
    """
    # 确保掩码和数据对齐
    if not df_data.equals(mask):
        # 这一步通常是为了确保传入的 df_data 和 mask 具有相同的形状、索引和列名
        # 在实际应用中，如果数据源保证了对齐，则可以跳过此检查。
        pass

    # 使用 numpy.where 基于 mask 的值来决定应用样式还是空字符串
    # 必须使用 .values 来处理底层的 numpy 数组，效率最高
    style_array = np.where(
        mask.values,
        f'background-color: {color}; font-weight: bold;',  # 更改过的值应用黄色背景和粗体
        ''  # 未更改的值返回空字符串（无样式）
    )

    # 样式函数必须返回一个具有相同索引和列的 DataFrame
    return pd.DataFrame(
        style_array,
        index=df_data.index,
        columns=df_data.columns
    )


def write_df(df, output_file_name, sheet_name: str = None):
    if sheet_name == "" or sheet_name is None:
        sheet_name = "修正数据表"

    excel_file = Path(output_file_name)
    if excel_file.exists():
        xls = pd.ExcelFile(excel_file)
        try:
            with pd.ExcelWriter(output_file_name, engine='openpyxl', mode='a',
                                if_sheet_exists='overlay') as writer:  # 注意：需要安装 openpyxl 库
                # if sheet_name not in xls.sheet_names:
                # 关键参数：
                # startrow=start_row: 指定从哪一行开始写入数据
                # header=header: 控制是否写入表头
                df.to_excel(
                    writer,
                    sheet_name=sheet_name,
                    index=False,  # 不写入 Pandas 索引
                    # startrow=start_row,
                    # header=header,
                )
        except PermissionError as e:
            logger.error(f"❌ 无法写入 {output_file_name}，请确认文件未被占用！{e}")
            raise e
        except Exception as e:
            logger.error(f"❌❌❌❌❌ Failed to write excel {e}")
            raise e

    else:
        df.to_excel(output_file_name, sheet_name=sheet_name, index=False)

    # 使用openpyxl加载刚才保存的Excel文件
    wb = load_workbook(excel_file)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        # 自适应调整列宽
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells if cell.value is not None)
            ws.column_dimensions[column_cells[0].column_letter].width = length + 2  # 可以根据需要调整额外的宽度

    # 保存调整后的Excel文件
    wb.save(excel_file)
    wb.close()


if __name__ == '__main__':
    """ 初始化配置 """
    _config = Configuration("config.yaml")

    # 输入输出文件路径
    input_file = _config.checker['input_file']
    output_log_file = _config.checker['output_log_file']
    output_adjusted_file = _config.checker['output_adjusted_file']

    # 保留小数位数
    ROUND_DECIMAL = _config.checker['ROUND_DECIMAL']

    # 调整比例阈值
    RATIO_MIN_PROVINCE = _config.ratio['RATIO_MIN_PROVINCE']
    RATIO_MAX_PROVINCE = _config.ratio['RATIO_MAX_PROVINCE']

    RATIO_MIN_1 = _config.ratio['RATIO_MIN_1']
    RATIO_MAX_1 = _config.ratio['RATIO_MAX_1']
    RATIO_MIN_2 = _config.ratio['RATIO_MIN_2']
    RATIO_MAX_2 = _config.ratio['RATIO_MAX_2']
    RATIO_MAX_3 = _config.ratio['RATIO_MAX_3']

    # --- 运行主程序 ---
    # try:
    df_2022, df_2017, df_national_ref, df_province_ref = load_data()
    output_file = process_data_for_all_crops_and_provinces(df_2022, df_2017, df_national_ref, df_province_ref)
    # except Exception as e:
    #     logger.error(f"程序执行失败: {e}")
